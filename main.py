from fastapi import FastAPI, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Annotated
from openai import OpenAI
from copy import copy
from openpyxl import load_workbook
import fitz
import base64
import os
import json
import io
import tempfile
import ezdxf

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))


@app.get("/")
def root():
    return {"status": "ok"}


def pdf_to_base64_image(pdf_bytes: bytes) -> str:
    pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = pdf[0]
    pix = page.get_pixmap(dpi=250)
    image_bytes = pix.tobytes("png")
    return base64.b64encode(image_bytes).decode("utf-8")


def extract_pdf_text(pdf_bytes: bytes) -> str:
    pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
    text = ""
    for page in pdf:
        text += page.get_text() + "\n"
    return text


def extract_from_pdf_image(base64_image: str, filename: str, pdf_text: str = ""):
    prompt = f"""
Extrae información estructurada de este plano de acero tipo Tekla.

Archivo fuente: {filename}

Texto extraído del PDF:
{pdf_text}

Reglas:
- Responde SOLO JSON válido.
- No inventes datos.
- Si un dato no aparece, usa null.
- Usa primero el texto extraído para MATERIAL LIST, pesos, proyecto, fase, fecha y archivo.
- Usa la imagen para validar y extraer cotas/barrenos.
- Extrae las posiciones horizontales de barrenaciones en milímetros.
- No cuentes los barrenos como resultado principal.
- No conviertas Ø 5/8" a milímetros; guárdalo como "5/8 in".
- La pestaña destino debe ser el perfil principal, por ejemplo "8C3014".
- Distingue entre el peso del material principal, clip y totales generales.
- Si existe clip L6, extrae sus datos por separado.
- Extrae Unit Weight Kg, Tot Weight Kg y Painting Area m2.
- No tomes como longitud principal valores de referencia como 6007 o 5950.
- La longitud principal está en MATERIAL LIST y en la etiqueta “perfil x longitud”.

Devuelve exactamente esta estructura:

{{
  "source_file": "{filename}",
  "mark": null,
  "cantidad_piezas": null,
  "target_sheet": null,
  "titulo": null,
  "project_no": null,
  "proyecto": null,
  "fase": null,
  "fecha": null,
  "revision": null,
  "archivo_plano": null,
  "material_principal": {{
    "mark": null,
    "qty": null,
    "profile": null,
    "grade": null,
    "length_mm": null,
    "unit_weight_kg": null,
    "total_weight_kg": null
  }},
  "clip": {{
    "mark": null,
    "qty": null,
    "profile": null,
    "grade": null,
    "length_mm": null,
    "unit_weight_kg": null,
    "total_weight_kg": null
  }},
  "totales": {{
    "unit_weight_kg": null,
    "total_weight_kg": null,
    "painting_area_m2": null
  }},
  "barrenos": {{
    "diametro": null,
    "posiciones_mm": [],
    "separacion_vertical_mm": null,
    "offset_vertical_mm": null
  }},
  "soldadura": {{
    "tipo": null,
    "tamano": null
  }},
  "elevacion": null,
  "validation_status": "pending"
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o",
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "system",
                "content": "Eres un extractor especializado en planos estructurales Tekla. Devuelves exclusivamente JSON válido."
            },
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{base64_image}"
                        }
                    }
                ]
            }
        ]
    )

    content = response.choices[0].message.content
    result = json.loads(content)

    material = result.get("material_principal") or {}

    if not result.get("mark"):
        result["mark"] = result.get("archivo_plano") or material.get("mark")

    if not result.get("target_sheet"):
        result["target_sheet"] = material.get("profile")

    return result


@app.post("/extract")
async def extract_pdf(file: UploadFile = File(...)):
    pdf_bytes = await file.read()
    base64_image = pdf_to_base64_image(pdf_bytes)
    pdf_text = extract_pdf_text(pdf_bytes)
    result = extract_from_pdf_image(base64_image, file.filename, pdf_text)
    return result


@app.post("/extract-batch")
async def extract_batch(
    files: Annotated[List[UploadFile], File(description="Archivos PDF a procesar")]
):
    items = []

    for file in files:
        pdf_bytes = await file.read()
        base64_image = pdf_to_base64_image(pdf_bytes)
        pdf_text = extract_pdf_text(pdf_bytes)
        result = extract_from_pdf_image(base64_image, file.filename, pdf_text)
        items.append(result)

    return {
        "items": items,
        "count": len(items)
    }


def extract_text_from_dxf(file_bytes: bytes, filename: str) -> dict:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    doc = ezdxf.readfile(tmp_path)
    msp = doc.modelspace()

    texts = []
    dimensions = []
    circles = []
    lines = []

    for entity in msp:
        entity_type = entity.dxftype()

        if entity_type in ["TEXT", "MTEXT"]:
            value = entity.plain_text() if entity_type == "MTEXT" else entity.dxf.text
            texts.append({
                "type": entity_type,
                "text": value,
                "layer": entity.dxf.layer
            })

        elif entity_type == "DIMENSION":
            dimensions.append({
                "type": entity_type,
                "layer": entity.dxf.layer,
                "text": getattr(entity.dxf, "text", None)
            })

        elif entity_type == "CIRCLE":
            circles.append({
                "type": entity_type,
                "layer": entity.dxf.layer,
                "center": [
                    float(entity.dxf.center.x),
                    float(entity.dxf.center.y),
                    float(entity.dxf.center.z)
                ],
                "radius": float(entity.dxf.radius)
            })

        elif entity_type == "LINE":
            lines.append({
                "type": entity_type,
                "layer": entity.dxf.layer,
                "start": [
                    float(entity.dxf.start.x),
                    float(entity.dxf.start.y),
                    float(entity.dxf.start.z)
                ],
                "end": [
                    float(entity.dxf.end.x),
                    float(entity.dxf.end.y),
                    float(entity.dxf.end.z)
                ]
            })

    return {
        "source_file": filename,
        "file_type": "dxf",
        "texts": texts,
        "dimensions": dimensions,
        "circles": circles,
        "lines_count": len(lines),
        "circles_count": len(circles),
        "texts_count": len(texts),
        "dimensions_count": len(dimensions)
    }


@app.post("/extract-cad")
async def extract_cad(file: UploadFile = File(...)):
    file_bytes = await file.read()
    filename = file.filename.lower()

    if filename.endswith(".dxf"):
        return extract_text_from_dxf(file_bytes, file.filename)

    if filename.endswith(".dwg"):
        return {
            "source_file": file.filename,
            "file_type": "dwg",
            "status": "conversion_required",
            "message": "DWG requiere conversión previa a DXF para esta etapa."
        }

    return {
        "error": "Formato no soportado. Usa PDF, DXF o DWG."
    }


COLUMN_MAP = {
    "default": {
        "folio": "B",
        "mark": "C",
        "cantidad": "D",
        "perfil": "E",
        "longitud": "F",
        "barrenos": ["G", "H", "I", "J", "K", "L", "M", "N", "O"],
        "nota": "P",
        "clip": "Q",
        "soldadura": "R",
        "simbolo": "S",
        "peso_unitario": "T",
        "zona": "U"
    }
}


def to_number(value):
    if value is None or value == "":
        return None
    try:
        num = float(value)
        return int(num) if num.is_integer() else num
    except Exception:
        return value


def get_records_from_payload(payload):
    if isinstance(payload, dict) and "items" in payload:
        return payload["items"]
    if isinstance(payload, list):
        return payload
    return [payload]


def normalize_record(record):
    material = record.get("material_principal") or {}
    totales = record.get("totales") or {}
    barrenos = record.get("barrenos") or {}
    clip = record.get("clip") or {}
    soldadura = record.get("soldadura") or {}

    mark = (
        record.get("mark")
        or record.get("archivo_plano")
        or material.get("mark")
        or ""
    )

    target_sheet = (
        record.get("target_sheet")
        or material.get("profile")
        or ""
    )

    return {
        "source_file": record.get("source_file"),
        "folio": record.get("folio") or "",
        "mark": mark,
        "cantidad": to_number(record.get("cantidad_piezas") or material.get("qty")),
        "perfil": material.get("profile") or target_sheet,
        "longitud": to_number(material.get("length_mm")),
        "target_sheet": target_sheet,
        "peso_unitario": to_number(material.get("unit_weight_kg") or totales.get("unit_weight_kg")),
        "peso_total": to_number(totales.get("total_weight_kg") or material.get("total_weight_kg")),
        "area_pintura": to_number(totales.get("painting_area_m2")),
        "barrenos": barrenos.get("posiciones_mm") or [],
        "diametro": barrenos.get("diametro"),
        "offset_vertical": to_number(barrenos.get("offset_vertical_mm")),
        "separacion_vertical": to_number(barrenos.get("separacion_vertical_mm")),
        "clip": clip,
        "soldadura": soldadura,
        "raw": record
    }


def copy_row_style(sheet, source_row, target_row):
    for col in range(1, sheet.max_column + 1):
        src = sheet.cell(row=source_row, column=col)
        dst = sheet.cell(row=target_row, column=col)

        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)


def find_insert_row(sheet):
    last_data_row = 1

    for row in range(1, sheet.max_row + 1):
        folio = sheet[f"B{row}"].value
        mark = sheet[f"C{row}"].value
        profile = sheet[f"E{row}"].value
        length = sheet[f"F{row}"].value

        if folio or mark or profile or length:
            last_data_row = row

    return last_data_row + 1


def write_record_to_sheet(sheet, record):
    data = normalize_record(record)
    colmap = COLUMN_MAP.get(data["target_sheet"], COLUMN_MAP["default"])

    insert_row = find_insert_row(sheet)

    sheet.insert_rows(insert_row)

    template_row = insert_row - 1
    if template_row >= 1:
        copy_row_style(sheet, template_row, insert_row)

    sheet[f'{colmap["folio"]}{insert_row}'] = data["folio"]
    sheet[f'{colmap["mark"]}{insert_row}'] = f'*{data["mark"]}' if data["mark"] else ""
    sheet[f'{colmap["cantidad"]}{insert_row}'] = data["cantidad"]
    sheet[f'{colmap["perfil"]}{insert_row}'] = data["perfil"]
    sheet[f'{colmap["longitud"]}{insert_row}'] = data["longitud"]

    holes = data["barrenos"]

    for i, col in enumerate(colmap["barrenos"]):
        sheet[f"{col}{insert_row}"] = holes[i] if i < len(holes) else None

    sheet[f'{colmap["nota"]}{insert_row}'] = "VER PLANO DE TALLER PARA HAB"

    clip = data["clip"] or {}
    sheet[f'{colmap["clip"]}{insert_row}'] = "CLIPS" if clip.get("mark") else ""

    soldadura = data["soldadura"] or {}
    sheet[f'{colmap["soldadura"]}{insert_row}'] = to_number(soldadura.get("tamano"))

    sheet[f'{colmap["simbolo"]}{insert_row}'] = "*"
    sheet[f'{colmap["peso_unitario"]}{insert_row}'] = data["peso_unitario"]
    sheet[f'{colmap["zona"]}{insert_row}'] = "S2"

    return {
        "mark": data["mark"],
        "target_sheet": data["target_sheet"],
        "row": insert_row,
        "holes": holes
    }


@app.post("/generate-excel")
async def generate_excel(
    template: UploadFile = File(...),
    records_json: str = File(...)
):
    template_bytes = await template.read()
    payload = json.loads(records_json)

    records = get_records_from_payload(payload)

    keep_vba = template.filename.lower().endswith(".xlsm")

    workbook = load_workbook(
        filename=io.BytesIO(template_bytes),
        keep_vba=keep_vba
    )

    inserted = []
    skipped = []

    for record in records:
        data = normalize_record(record)
        target_sheet = data["target_sheet"]

        if not target_sheet:
            skipped.append({
                "source_file": data.get("source_file"),
                "reason": "No tiene target_sheet"
            })
            continue

        if target_sheet not in workbook.sheetnames:
            skipped.append({
                "source_file": data.get("source_file"),
                "mark": data.get("mark"),
                "target_sheet": target_sheet,
                "reason": "No existe la hoja destino en el Excel"
            })
            continue

        sheet = workbook[target_sheet]
        result = write_record_to_sheet(sheet, record)
        inserted.append(result)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = "planos_a_excel_resultado.xlsm" if keep_vba else "planos_a_excel_resultado.xlsx"

    media_type = (
        "application/vnd.ms-excel.sheet.macroEnabled.12"
        if keep_vba
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-Inserted-Rows": json.dumps(inserted),
        "X-Skipped-Rows": json.dumps(skipped)
    }

    return StreamingResponse(
        output,
        media_type=media_type,
        headers=headers
    )